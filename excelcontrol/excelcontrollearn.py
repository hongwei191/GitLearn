# coding: utf-8
import openpyxl
from openpyxl import load_workbook
import xlrd

def openyxlcontrolexcel(path):
    wb =  load_workbook(path,read_only=False)
    wbactive= wb.active
    sheet =wb['feature_index']
    rows = sheet.max_row
    cols=sheet.max_column
    print(rows,cols)

def xrldcontrolexcel(path):
    table = xlrd.open_workbook(path)
    sheet = table.sheet_by_index(0)
    rows= sheet.nrows
    cols = sheet.ncols
    print(rows,cols)
    data=[]
    valuestr = ''
    for i in range(rows):
        list_value = []
        for j in range(cols):
            value= sheet.cell(i,j).value
            value = valuestr+str(value)
            list_value.append(value)
        data.append(list_value)
    print(list_value)
    print(data)



if __name__ == '__main__':
    money = r'‪C:\Users\94599\Desktop\房贷还款计划.xlsx'
    path = r'F:\NvAutoCheck\Featureindex.xlsx'
    openyxlcontrolexcel(path)
    # xrldcontrolexcel(path)